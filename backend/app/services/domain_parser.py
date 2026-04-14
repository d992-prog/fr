from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from urllib.parse import urlsplit

from fastapi import UploadFile
from openpyxl import load_workbook

DOMAIN_CANDIDATE_PATTERN = re.compile(r"(?<![\w-])(?:[\w-]+(?:\.[\w-]+)+)(?![\w-])", re.UNICODE)
ASCII_LABEL_PATTERN = re.compile(r"^[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?$")
ALLOWED_EXTENSIONS = {".txt", ".csv", ".xlsx"}


def normalize_domain(value: str) -> str | None:
    candidate = value.strip().lower().strip("\"'")
    if not candidate:
        return None

    if "://" in candidate:
        parsed = urlsplit(candidate)
        candidate = parsed.hostname or ""
    else:
        candidate = candidate.split("/", 1)[0].split("?", 1)[0].split("#", 1)[0]
        if "@" in candidate:
            candidate = candidate.rsplit("@", 1)[-1]
        if ":" in candidate and candidate.rsplit(":", 1)[-1].isdigit():
            candidate = candidate.rsplit(":", 1)[0]

    candidate = candidate.strip(".")
    if not candidate:
        return None

    try:
        ascii_domain = candidate.encode("idna").decode("ascii").lower()
    except UnicodeError:
        return None

    labels = ascii_domain.split(".")
    if len(labels) < 2:
        return None

    for label in labels:
        if not label or len(label) > 63 or not ASCII_LABEL_PATTERN.fullmatch(label):
            return None

    zone = labels[-1]
    if len(zone) < 2 or zone.isdigit():
        return None

    return ".".join(labels)


def extract_domains_from_text(content: str) -> list[str]:
    found: set[str] = set()
    for match in DOMAIN_CANDIDATE_PATTERN.finditer(content.lower()):
        normalized = normalize_domain(match.group(0))
        if normalized:
            found.add(normalized)
    return sorted(found)


def extract_domains_from_rows(rows: list[list[str]]) -> list[str]:
    found: set[str] = set()
    for row in rows:
        for cell in row:
            for match in DOMAIN_CANDIDATE_PATTERN.finditer(cell.lower()):
                normalized = normalize_domain(match.group(0))
                if normalized:
                    found.add(normalized)
    return sorted(found)


async def parse_upload(file: UploadFile, max_bytes: int) -> list[str]:
    extension = Path(file.filename or "").suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError("Unsupported file type. Allowed: .txt, .csv, .xlsx")

    raw = await file.read()
    if len(raw) > max_bytes:
        raise ValueError("Uploaded file is too large")

    if extension == ".txt":
        return extract_domains_from_text(raw.decode("utf-8", errors="ignore"))

    if extension == ".csv":
        text = raw.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = [[cell.strip() for cell in row] for row in reader]
        return extract_domains_from_rows(rows)

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    rows: list[list[str]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell or "").strip() for cell in row])
    return extract_domains_from_rows(rows)
